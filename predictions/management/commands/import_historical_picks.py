import hashlib
from decimal import Decimal, InvalidOperation
from datetime import datetime, date, time

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from predictions.models import HistoricalPick


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_decimal(value):
    if value is None or value == "":
        return None

    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def clean_date(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def clean_time(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.time()

    if isinstance(value, time):
        return value

    return None


def clean_outcome(value):
    value = clean_text(value).upper()

    if value in ["W", "WIN", "WON"]:
        return "W"

    if value in ["L", "LOSS", "LOST"]:
        return "L"

    if value in ["R", "REFUND", "VOID", "PUSH"]:
        return "R"

    if value in ["P", "PENDING"]:
        return "P"

    return "U"


def make_hash(row_values):
    raw = "|".join(clean_text(value) for value in row_values)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


class Command(BaseCommand):
    help = "Import historical MMA picks from the existing Google Sheets Excel file."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)
        parser.add_argument(
            "--promotion",
            type=str,
            default="UFC",
            help="Promotion name for this spreadsheet, e.g. UFC, PFL, OKTAGON.",
    )

    def handle(self, *args, **options):
        file_path = options["promotion"].strip().upper()

        try:
            workbook = load_workbook(file_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {file_path}")

        if "Picks" not in workbook.sheetnames:
            raise CommandError("Could not find a sheet named 'Picks'.")

        sheet = workbook["Picks"]

        headers = {}
        for cell in sheet[1]:
            if cell.value:
                headers[str(cell.value).strip().lower()] = cell.column

        required_columns = [
            "date",
            "time",
            "country",
            "event",
            "fight",
            "bet",
            "type",
            "stake",
            "odds",
            "outcome",
            "p/l",
            "total p/l",
        ]

        missing = [column for column in required_columns if column not in headers]
        if missing:
            raise CommandError(f"Missing columns: {', '.join(missing)}")

        created_count = 0
        skipped_count = 0

        for row_number in range(2, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_number, column=col).value for col in range(1, 13)]

            if not any(row_values):
                continue

            source_hash = make_hash(row_values)

            if HistoricalPick.objects.filter(source_hash=source_hash).exists():
                skipped_count += 1
                continue

            event_name = clean_text(sheet.cell(row=row_number, column=headers["event"]).value)
            fight_name = clean_text(sheet.cell(row=row_number, column=headers["fight"]).value)
            pick_name = clean_text(sheet.cell(row=row_number, column=headers["bet"]).value)

            if not event_name and not fight_name and not pick_name:
                continue

            HistoricalPick.objects.create(
                date=clean_date(sheet.cell(row=row_number, column=headers["date"]).value),
                time=clean_time(sheet.cell(row=row_number, column=headers["time"]).value),
                promotion=promotion,
                country=clean_text(sheet.cell(row=row_number, column=headers["country"]).value),
                event_name=event_name,
                fight_name=fight_name,
                pick_name=pick_name,
                bet_type=clean_text(sheet.cell(row=row_number, column=headers["type"]).value),
                stake=clean_decimal(sheet.cell(row=row_number, column=headers["stake"]).value),
                odds=clean_decimal(sheet.cell(row=row_number, column=headers["odds"]).value),
                outcome=clean_outcome(sheet.cell(row=row_number, column=headers["outcome"]).value),
                profit_loss=clean_decimal(sheet.cell(row=row_number, column=headers["p/l"]).value),
                sheet_total_profit_loss=clean_decimal(sheet.cell(row=row_number, column=headers["total p/l"]).value),
                source_row=row_number,
                source_hash=source_hash,
            )

            created_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete. Created {created_count} picks. Skipped {skipped_count} duplicates."
            )
        )
